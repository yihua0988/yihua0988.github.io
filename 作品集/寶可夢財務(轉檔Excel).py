"""
請洗資料並轉換幣別存入Excel（只保留2020~2024年資料）
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font

# 假設匯率 1 USD = 135 JPY
exchange_rate = 135

# 年份 2020~2024
years = ["2020","2021","2022","2023","2024"]

# 類別
categories = ["營業額", "營業利益", "經常利益", "純利益", "股東權益", "總負債", "保留盈餘"]

# 以日圓為單位(資料內容)，只取 2020~2024
data_in_yen = [
    [120019, 204209, 234528, 297508, 410932],  # 營業額
    [27804, 59860, 66645, 88692, 100749],      # 營業利益
    [27970, 62241, 69105, 90836, 101487],      # 經常利益
    [18630, 41392, 48854, 62701, 70343],       # 純利益
    [79842, 121234, 170357, 224081, 294425],   # 股東權益
    [25406, 57323, 58241, 17942, 111930],      # 總負債
    [79545, 120937, 169792, 223516, 293860],   # 保留盈餘
]

# 轉換為億 USD，缺失值用 "-"
data_in_usd = []
for row in data_in_yen:
    data_in_usd.append([
        round(value * 1000000 / exchange_rate / 100000000, 2) if value is not None else '-' 
        for value in row
    ])

# Excel 檔案路徑（程式資料夾）
script_folder = os.path.dirname(os.path.abspath(__file__))
xlsx_filename = os.path.join(script_folder, 'pokemon_financial_2020_2024.xlsx')

# 建立工作簿
wb = Workbook()
ws = wb.active
ws.title = "財務資料"

# 設定框線
thin_side = Side(style='thin')
thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

# 設定字型
default_font = Font(name='標楷體', size=14)
bold_font = Font(name='標楷體', size=14, bold=True)

# 寫入標題
ws.cell(row=1, column=1, value="西元年").border = thin_border
ws.cell(row=1, column=1).font = bold_font
ws.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')

for col, year in enumerate(years, start=2):
    ws.cell(row=1, column=col, value=year).border = thin_border
    ws.cell(row=1, column=col).font = default_font
    ws.cell(row=1, column=col).alignment = Alignment(horizontal='center', vertical='center')

# 寫入科目與數字
for row_idx, category in enumerate(categories, start=2):
    ws.cell(row=row_idx, column=1, value=category).border = thin_border
    ws.cell(row=row_idx, column=1).font = bold_font
    ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal='center', vertical='center')
    
    for col_idx, value in enumerate(data_in_usd[row_idx-2], start=2):
        ws.cell(row=row_idx, column=col_idx, value=value).border = thin_border
        ws.cell(row=row_idx, column=col_idx).font = default_font
        ws.cell(row=row_idx, column=col_idx).alignment = Alignment(horizontal='center', vertical='center')

# 最下面一列寫單位
unit_row = len(categories) + 2
ws.cell(row=unit_row, column=1, value="單位：億 USD").border = thin_border
ws.cell(row=unit_row, column=1).font = bold_font
ws.cell(row=unit_row, column=1).alignment = Alignment(horizontal='center', vertical='center')
for col in range(2, len(years)+2):
    ws.cell(row=unit_row, column=col).border = thin_border
    ws.cell(row=unit_row, column=col).font = default_font
    ws.cell(row=unit_row, column=col).alignment = Alignment(horizontal='center', vertical='center')

# 自動調整欄寬
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        if cell.value:
            max_length = max(max_length, len(str(cell.value)))
    if column == 'A':
        ws.column_dimensions[column].width = int(max_length * 2.5)
    else:
        ws.column_dimensions[column].width = max_length + 4

# 儲存 Excel
wb.save(xlsx_filename)
print(f"Excel 文件已成功創建，存放在: {xlsx_filename}")
